import os
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse
import json
import re
import datetime
import openpyxl
import pdfplumber
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Sum, Count, Q
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.utils import timezone

from .models import (
    OrdenCompra, FMR, Entrega, Costo, Trazabilidad, ItemOC, PackingListItem, Factura,
    CostoMaterial, CostoManoObra, Cargo, ManoDeObra, MateriaPrima, PackingList,
    Cotizacion, ItemCotizacion, GuiaDespacho, ItemGuia,
)
from .forms import (
    OrdenCompraForm, OrdenCompraEditForm, FMRForm, EntregaForm, CostoForm, ItemOCForm,
    PackingListItemForm, FacturaForm, CostoMaterialForm, CostoManoObraForm,
    MateriaPrimaForm, ManoDeObraForm, PackingListForm,
    CotizacionForm, ItemCotizacionForm, GuiaDespachoForm, ItemGuiaForm,
)


# ──────────────────────────────────────────────────────────────────────────────
# AUTH
# ──────────────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('project_list')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect(request.GET.get('next', 'project_list'))
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'core/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ──────────────────────────────────────────────────────────────────────────────
# DASHBOARD (KPI overview — accessed from navbar)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    from datetime import date
    from dateutil.relativedelta import relativedelta

    # ── Totales globales ─────────────────────────────────────────────────────
    total_ocs       = OrdenCompra.objects.count()
    total_revenue   = OrdenCompra.objects.aggregate(t=Sum('valor_total'))['t'] or Decimal('0.00')
    
    # Sumar costos generales, materiales de compras detalladas y mano de obra
    total_costs_general = Costo.objects.aggregate(t=Sum('monto'))['t'] or Decimal('0.00')
    total_materials = sum(mc.total for mc in CostoMaterial.objects.all())
    total_mano_obra = sum(moc.total for moc in CostoManoObra.objects.all())
    
    total_costs     = total_costs_general + total_materials + total_mano_obra
    global_margin   = total_revenue - total_costs
    global_margin_pct = float(global_margin / total_revenue * 100) if total_revenue > 0 else 0

    # ── OC por estado ────────────────────────────────────────────────────────
    status_counts = (
        OrdenCompra.objects.values('estado')
        .annotate(count=Count('numero_oc'))
        .order_by('estado')
    )
    status_dict = {item['estado']: item['count'] for item in status_counts}

    # ── Semaforos: conteo por color ──────────────────────────────────────────
    # No se pueden hacer con annotate puro (son @property), pero se calculan
    # en Python sobre el queryset ya cargado — es un conjunto pequeño.
    todas_ocs = OrdenCompra.objects.only(
        'numero_oc', 'estado', 'fecha_compromiso', 'porcentaje_entregado'
    )
    semaforos = {'plazo': {'green':0,'yellow':0,'red':0,'grey':0},
                 'avance': {'green':0,'yellow':0,'red':0,'grey':0},
                 'docs': {'green':0,'yellow':0,'red':0,'grey':0}}
    for oc in todas_ocs:
        semaforos['plazo'][oc.semaforo_plazo]   += 1
        semaforos['avance'][oc.semaforo_avance] += 1
    # docs requiere prefetch de fmrs; lo calculamos sobre las primeras 200
    ocs_docs = OrdenCompra.objects.prefetch_related('fmrs').only(
        'numero_oc','oc_link','oc_file','plano_link','plano_file','fmr_link','fmr_file'
    )[:200]
    for oc in ocs_docs:
        semaforos['docs'][oc.semaforo_docs] += 1

    # ── Monto pendiente de facturar ──────────────────────────────────────────
    # OC sin facturas o con todas sus facturas en estado 'pendiente'
    ocs_sin_factura = OrdenCompra.objects.filter(
        facturas__isnull=True
    ).aggregate(t=Sum('valor_total'))['t'] or Decimal('0.00')
    facturas_pendientes_monto = (
        Factura.objects.filter(estado='pendiente')
        .aggregate(t=Sum('monto'))['t'] or Decimal('0.00')
    )
    monto_pendiente_facturar = ocs_sin_factura + facturas_pendientes_monto

    # ── Top clientes por monto total ─────────────────────────────────────────
    top_clientes_monto = (
        OrdenCompra.objects
        .values('cliente')
        .annotate(total_monto=Sum('valor_total'), total_ocs=Count('numero_oc'))
        .order_by('-total_monto')[:8]
    )

    # ── Top clientes por margen (calculado en Python, top 8 ya cargados) ─────
    # Unimos costos por cliente para calcular el margen preciso
    clientes_margen = []
    for row in top_clientes_monto:
        cliente_name = row['cliente']
        costos_general = (
            Costo.objects
            .filter(orden_compra__cliente=cliente_name)
            .aggregate(t=Sum('monto'))['t'] or Decimal('0.00')
        )
        costos_mat = sum(
            mc.total for mc in CostoMaterial.objects.filter(orden_compra__cliente=cliente_name)
        )
        costos_mo = sum(
            moc.total for moc in CostoManoObra.objects.filter(orden_compra__cliente=cliente_name)
        )
        
        costos_cliente = costos_general + costos_mat + costos_mo
        monto = row['total_monto'] or Decimal('0.00')
        margen = monto - costos_cliente
        margen_pct = float(margen / monto * 100) if monto > 0 else 0
        clientes_margen.append({
            'cliente': cliente_name,
            'total_monto': monto,
            'total_ocs': row['total_ocs'],
            'margen': margen,
            'margen_pct': round(margen_pct, 1),
        })
    clientes_margen.sort(key=lambda x: x['margen'], reverse=True)

    # ── Reporte mensual ultimos 6 meses ──────────────────────────────────────
    hoy = date.today()
    meses = []
    for i in range(5, -1, -1):
        inicio = (hoy - relativedelta(months=i)).replace(day=1)
        fin    = (inicio + relativedelta(months=1))
        ocs_mes = OrdenCompra.objects.filter(
            fecha_oc__gte=inicio, fecha_oc__lt=fin
        ).aggregate(
            cantidad=Count('numero_oc'),
            monto=Sum('valor_total')
        )
        facturas_mes = Factura.objects.filter(
            fecha_emision__gte=inicio, fecha_emision__lt=fin
        ).aggregate(facturado=Sum('monto'))
        meses.append({
            'mes': inicio.strftime('%b %Y'),
            'cantidad_ocs': ocs_mes['cantidad'] or 0,
            'monto_ocs': ocs_mes['monto'] or Decimal('0.00'),
            'facturado': facturas_mes['facturado'] or Decimal('0.00'),
        })

    # ── Listas de detalle ────────────────────────────────────────────────────
    recent_deliveries = Entrega.objects.select_related('orden_compra').order_by('-fecha_entrega')[:6]
    pending_ocs       = OrdenCompra.objects.filter(
        estado__in=['En proceso','Pendiente']
    ).order_by('fecha_compromiso')[:6]
    cost_by_category  = Costo.objects.values('categoria').annotate(total=Sum('monto')).order_by('-total')

    context = {
        'total_ocs': total_ocs,
        'total_revenue': total_revenue,
        'total_costs': total_costs,
        'global_margin': global_margin,
        'global_margin_pct': round(global_margin_pct, 2),
        'status_dict': status_dict,
        'semaforos': semaforos,
        'monto_pendiente_facturar': monto_pendiente_facturar,
        'clientes_margen': clientes_margen,
        'reporte_meses': meses,
        'recent_deliveries': recent_deliveries,
        'pending_ocs': pending_ocs,
        'cost_by_category': cost_by_category,
    }
    return render(request, 'core/dashboard.html', context)


# ──────────────────────────────────────────────────────────────────────────────
# PROJECT LIST — Main screen
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def project_list(request):
    queryset = OrdenCompra.objects.all().prefetch_related('fmrs', 'entregas')

    # Filters
    q = request.GET.get('q', '').strip()
    estado_filter = request.GET.get('estado', '')

    if q:
        queryset = queryset.filter(
            Q(numero_oc__icontains=q) |
            Q(cliente__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(proyecto__icontains=q) |
            Q(guia_despacho_resumen__icontains=q) |
            Q(factura_resumen__icontains=q) |
            Q(fmrs__fmr_code__icontains=q) |
            Q(facturas__numero_factura__icontains=q)  # busqueda en modelo Factura
        ).distinct()

    if estado_filter:
        queryset = queryset.filter(estado=estado_filter)

    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    estados = OrdenCompra.ESTADOS

    context = {
        'page_obj': page_obj,
        'q': q,
        'estado_filter': estado_filter,
        'estados': estados,
        'total_count': queryset.count(),
    }
    return render(request, 'core/project_list.html', context)


# ──────────────────────────────────────────────────────────────────────────────
# SEARCH
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def search(request):
    query = request.GET.get('q', '').strip()
    results_oc = []
    results_fmr = []
    results_cotizacion = []
    results_guia = []

    if query:
        results_oc = OrdenCompra.objects.filter(
            Q(numero_oc__icontains=query) |
            Q(cliente__icontains=query) |
            Q(proyecto__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(guia_despacho_resumen__icontains=query) |
            Q(factura_resumen__icontains=query) |
            Q(cotizaciones__numero_cotizacion__icontains=query)
        ).distinct()

        results_fmr = FMR.objects.filter(
            Q(fmr_code__icontains=query) |
            Q(guia_despacho__icontains=query) |
            Q(factura__icontains=query) |
            Q(cotizacion__icontains=query)
        ).distinct()

        results_cotizacion = Cotizacion.objects.filter(
            Q(numero_cotizacion__icontains=query) |
            Q(razon_social__icontains=query) |
            Q(cliente_id__icontains=query)
        ).distinct()

        results_guia = GuiaDespacho.objects.filter(
            Q(numero_guia__icontains=query) |
            Q(receptor_nombre__icontains=query)
        ).distinct()

        if (len(results_oc) == 1 and len(results_fmr) == 0
                and len(results_cotizacion) == 0 and len(results_guia) == 0):
            return redirect('project_detail', numero_oc=results_oc[0].numero_oc)
        elif len(results_fmr) == 1 and len(results_oc) == 0:
            if results_fmr[0].orden_compra:
                return redirect('project_detail', numero_oc=results_fmr[0].orden_compra.numero_oc)

    context = {
        'query': query,
        'results_oc': results_oc,
        'results_fmr': results_fmr,
        'results_cotizacion': results_cotizacion,
        'results_guia': results_guia,
    }
    return render(request, 'core/search_results.html', context)


def registrar_trazabilidad(orden_compra, accion, detalle, usuario):
    Trazabilidad.objects.create(
        orden_compra=orden_compra,
        usuario=usuario if usuario and usuario.is_authenticated else None,
        accion=accion,
        detalle=detalle
    )

# ──────────────────────────────────────────────────────────────────────────────
# OC CRUD
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def oc_create(request):
    if request.method == 'POST':
        form = OrdenCompraForm(request.POST, request.FILES)
        if form.is_valid():
            oc = form.save()
            registrar_trazabilidad(
                oc, 
                "Creación de OC", 
                f"Orden de Compra {oc.numero_oc} registrada en el sistema.", 
                request.user
            )
            
            # Auto-save de ítems BOM extraídos del PDF
            items_json = request.POST.get('items_extraidos_json')
            import traceback
            try:
                with open("debug_items_import.txt", "w") as dbg_file:
                    dbg_file.write(f"POST received. Has items_extraidos_json: {bool(items_json)}\n")
                    dbg_file.write(f"Length of JSON String: {len(items_json) if items_json else 0}\n")
                    if items_json:
                        dbg_file.write(f"JSON Content preview: {items_json[:200]}\n")
            except:
                pass

            if items_json:
                import json
                try:
                    items_list = json.loads(items_json)
                    with open("debug_items_import.txt", "a") as dbg_file:
                        dbg_file.write(f"JSON loaded! Count: {len(items_list)}\n")
                    for data in items_list:
                        # Clean numbers safely
                        try:
                            c = str(data.get('cantidad', '0')).replace(',', '').strip()
                            c = float(c) if c else 0.0
                        except: c = 0.0
                        
                        try:
                            pu = str(data.get('precio_unitario', '0')).replace(',', '').strip()
                            pu = float(pu) if pu else 0.0
                        except: pu = 0.0
                        
                        try:
                            pt = str(data.get('precio_total', '0')).replace(',', '').strip()
                            pt = float(pt) if pt else 0.0
                        except: pt = 0.0

                        try:
                            # Try to parse e.g., '08Jul26'
                            fe_str = data.get('fecha_entrega', '')
                            if fe_str:
                                import datetime
                                d = datetime.datetime.strptime(fe_str, "%d%b%y")
                                fe = d.strftime("%Y-%m-%d")
                            else:
                                fe = None
                        except:
                            fe = None

                        try:
                            # Try parse peso_unitario_kg if extracted
                            peso_str = str(data.get('peso_unitario_kg', '')).strip()
                            try:
                                peso_val = float(peso_str) if peso_str else None
                            except (ValueError, TypeError):
                                peso_val = None

                            ItemOC.objects.create(
                                orden_compra=oc,
                                linea=str(data.get('linea', 1))[:50],
                                item_code=str(data.get('item_code', ''))[:100],
                                codigo=str(data.get('item_code', ''))[:100], # Guardamos codigo plano como item_code
                                size_code=str(data.get('size_code', ''))[:100],
                                descripcion=str(data.get('descripcion', ''))[:255],
                                cantidad=c,
                                unidad=str(data.get('uom', 'EA'))[:20],
                                uom=str(data.get('uom', 'EA'))[:100],
                                peso_unitario_kg=peso_val,
                                precio_unitario=pu,
                                precio_total=pt,
                                fecha_entrega=fe
                            )
                            with open("debug_items_import.txt", "a") as dbg_file:
                                dbg_file.write(f"Success creating item: {data.get('item_code', '')}\n")
                        except Exception as inner_e:
                            with open("debug_items_import.txt", "a") as dbg_file:
                                dbg_file.write(f"ERROR creating specific item {data.get('item_code', '')}: {str(inner_e)}\n")

                    if items_list:
                        messages.success(request, f'Se cargaron automáticamente {len(items_list)} ítems desde el documento.')
                except Exception as e:
                    with open("debug_items_import.txt", "a") as dbg_file:
                        dbg_file.write(f"ERROR parsing items JSON or loop: {str(e)}\n{traceback.format_exc()}\n")
                    messages.warning(request, f'Hubo un problema cargando los materiales automáticamente: {e}')

            messages.success(request, f'Orden de Compra {oc.numero_oc} creada exitosamente.')
            return redirect('project_detail', numero_oc=oc.numero_oc)
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = OrdenCompraForm()

    return render(request, 'core/oc_form.html', {'form': form, 'titulo': 'Nueva Orden de Compra', 'is_create': True})


@login_required
def oc_edit(request, numero_oc):
    project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    if request.method == 'POST':
        form = OrdenCompraEditForm(request.POST, request.FILES, instance=project)
        if form.is_valid():
            form.save()
            registrar_trazabilidad(
                project, 
                "Modificación de OC", 
                "Se actualizaron los datos o el respaldo documental de la Orden de Compra.", 
                request.user
            )
            messages.success(request, 'Proyecto actualizado correctamente.')
            return redirect('project_detail', numero_oc=numero_oc)
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = OrdenCompraEditForm(instance=project)

    return render(request, 'core/oc_form.html', {
        'form': form, 'project': project,
        'titulo': f'Editar OC: {numero_oc}', 'is_create': False
    })


@login_required
def oc_delete(request, numero_oc):
    project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    if request.method == 'POST':
        numero = project.numero_oc
        registrar_trazabilidad(
            None, 
            "Eliminación de OC", 
            f"Se eliminó permanentemente la Orden de Compra {numero} y todos sus registros asociados.", 
            request.user
        )
        project.delete()
        messages.success(request, f'Orden de Compra {numero} eliminada exitosamente.')
        return redirect('project_list')

    # Para GET request, mostramos la página de confirmación
    entregas_count = project.entregas.count()
    facturas_count = project.facturas.count()
    packing_lists_count = project.packing_lists.count()

    context = {
        'project': project,
        'entregas_count': entregas_count,
        'facturas_count': facturas_count,
        'packing_lists_count': packing_lists_count,
    }
    return render(request, 'core/oc_confirm_delete.html', context)


# ──────────────────────────────────────────────────────────────────────────────
# PROJECT DETAIL
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def project_detail(request, numero_oc):
    project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)

    # ── CARGAR COSTOS TRADICIONALES Y DETALLADOS ─────────────────────────────
    costs = project.costos.all().order_by('-fecha')
    costs_general_total = costs.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')

    # Costos detallados de materiales
    material_costs = project.materias_primas.all()
    material_total = project.costo_total_materiales

    # Costos detallados de mano de obra
    mano_obra_costs = project.manos_de_obra.select_related('cargo').all()
    mano_obra_total = project.costo_total_mano_obra

    # Gran total costos (general + materiales + mano de obra)
    costs_total = costs_general_total + project.costo_total_trabajo

    budget = project.valor_total or Decimal('0.00')
    margin = project.utilidad_real
    margin_pct = project.porcentaje_utilidad
    budget_used_pct = float(costs_total / budget * 100) if budget > 0 else 0

    fmrs = project.fmrs.all()
    deliveries = project.entregas.prefetch_related('packing_list_items__item_oc').all().order_by('-fecha_entrega')
    items = project.items.all().order_by('linea')
    trazabilidades = project.trazabilidades.all().order_by('-fecha_hora')

    # ── CALCULOS POR KILO (BOM) ──────────────────────────────────────────────
    total_weight = project.peso_total_kg
    items_total_value = sum(item.valor_total for item in items)

    # Cálculos detallados por kilo
    costo_por_kilo = project.costo_por_kilo
    utilidad_por_kilo = project.utilidad_por_kilo

    # Formularios
    entrega_form = EntregaForm()
    fmr_form = FMRForm()
    costo_form = CostoForm(initial={'fecha': timezone.now().date()})
    item_form = ItemOCForm()
    packing_list_form = PackingListItemForm(orden_compra=project)
    
    # Nuevos formularios detallados (asociados a nuevos modelos MateriaPrima y ManoDeObra)
    costo_material_form = MateriaPrimaForm()
    costo_mano_obra_form = ManoDeObraForm()

    context = {
        'project': project,
        'costs': costs,
        'material_costs': material_costs,
        'mano_obra_costs': mano_obra_costs,
        
        # Totales agregados
        'costs_general_total': costs_general_total,
        'material_total': material_total,
        'mano_obra_total': mano_obra_total,
        'costs_total': costs_total,
        'margin': margin,
        'margin_pct': round(margin_pct, 2),
        'budget_used_pct': min(round(budget_used_pct, 2), 100),
        
        'fmrs': fmrs,
        'deliveries': deliveries,
        'items': items,
        'trazabilidades': trazabilidades,
        'total_weight': total_weight,
        'items_total_value': items_total_value,
        
        # Indicadores por Kilo
        'costo_por_kilo': round(costo_por_kilo, 2),
        'utilidad_por_kilo': round(utilidad_por_kilo, 2),
        
        # Formularios
        'entrega_form': entrega_form,
        'fmr_form': fmr_form,
        'costo_form': costo_form,
        'item_form': item_form,
        'packing_list_form': packing_list_form,
        'costo_material_form': costo_material_form,
        'costo_mano_obra_form': costo_mano_obra_form,
        'dias_restantes': project.dias_restantes_calculado,
    }
    return render(request, 'core/project_detail.html', context)


# ──────────────────────────────────────────────────────────────────────────────
# INLINE ACTIONS (POST only — redirect back to detail)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def add_cost(request, numero_oc):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        form = CostoForm(request.POST)
        if form.is_valid():
            costo = form.save(commit=False)
            costo.orden_compra = project
            costo.save()
            registrar_trazabilidad(
                project,
                "Registro de Costo",
                f"Se registró un gasto en categoría '{costo.categoria}' por ${costo.monto:,.0f}. Detalle: {costo.descripcion} (Ref: {costo.documento_referencia or 'Sin ref'})",
                request.user
            )
            messages.success(request, f'Gasto de ${costo.monto:,.0f} registrado en Centro de Costos.')
        else:
            messages.error(request, f'Error al registrar gasto: {form.errors}')
    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def add_entrega(request, numero_oc):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        form = EntregaForm(request.POST, request.FILES)
        if form.is_valid():
            entrega = form.save(commit=False)
            entrega.orden_compra = project
            entrega.save()
            # Recalculate % delivered
            project.recalcular_porcentaje()
            archivo_str = " con archivo adjunto" if entrega.guia_file else ""
            registrar_trazabilidad(
                project,
                "Despacho Parcial",
                f"Se registró despacho con Guía N° {entrega.guia_despacho} ({entrega.get_estado_display()}){archivo_str}. Detalle: {entrega.cantidad_entregada}",
                request.user
            )
            messages.success(request, f'Guía {entrega.guia_despacho} registrada exitosamente.')
        else:
            messages.error(request, f'Error al registrar entrega: {form.errors}')
    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def add_fmr(request, numero_oc):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        form = FMRForm(request.POST, request.FILES)
        if form.is_valid():
            fmr = form.save(commit=False)
            fmr.orden_compra = project
            fmr.save()
            archivo_str = " con archivo adjunto" if fmr.registro_file else ""
            registrar_trazabilidad(
                project,
                "Vinculación de FMR",
                f"Se vinculó el registro FMR N° {fmr.fmr_code}{archivo_str} (Cotización: {fmr.cotizacion or 'Sin cotización'})",
                request.user
            )
            messages.success(request, f'FMR {fmr.fmr_code} vinculado al proyecto.')
        else:
            messages.error(request, f'Error al crear FMR: {form.errors}')
    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def add_item(request, numero_oc):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        form = ItemOCForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.orden_compra = project
            item.save()
            registrar_trazabilidad(
                project,
                "Item Agregado",
                f"Se agregó el item {item.linea}: {item.descripcion} (Cantidad: {item.cantidad} {item.unidad}, Precio Unitario: ${item.precio_unitario:,.0f})",
                request.user
            )
            messages.success(request, f'Item {item.linea} agregado exitosamente.')
        else:
            messages.error(request, f'Error al agregar item: {form.errors}')
    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def edit_item(request, numero_oc, item_id):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        item = get_object_or_404(ItemOC, id=item_id, orden_compra=project)
        
        form = ItemOCForm(request.POST, instance=item)
        if form.is_valid():
            nueva_cantidad = form.cleaned_data['cantidad']
            nueva_entregada = form.cleaned_data['cantidad_entregada']
            
            if nueva_cantidad < nueva_entregada:
                messages.error(request, f'No puedes reducir la cantidad ({nueva_cantidad}) por debajo de lo entregado ({nueva_entregada} unidades).')
                return redirect('project_detail', numero_oc=numero_oc)
            
            cambios = []
            for field in form.changed_data:
                cambios.append(f"{field.capitalize()} a '{form.cleaned_data[field]}'")
                
            item = form.save()
            project.recalcular_porcentaje()
            
            str_cambios = ", ".join(cambios) if cambios else "Sin cambios"
            registrar_trazabilidad(
                project,
                "Edición de Ítem BOM",
                f"Se modificó el ítem {item.linea} ({item.descripcion}). Cambios: {str_cambios}.",
                request.user
            )
            messages.success(request, f'Ítem {item.linea} actualizado exitosamente.')
        else:
            messages.error(request, f'Error al actualizar el ítem: {form.errors}')
    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def cost_center_overview(request):
    projects_with_costs = OrdenCompra.objects.annotate(
        costs_total=Sum('costos__monto')
    ).order_by('-valor_total')

    summary_list = []
    total_budget = Decimal('0.00')
    total_expenses = Decimal('0.00')

    for p in projects_with_costs:
        budget = p.valor_total or Decimal('0.00')
        expenses = p.costs_total or Decimal('0.00')
        margin = budget - expenses
        margin_pct = float(margin / budget * 100) if budget > 0 else 0
        summary_list.append({
            'project': p, 'budget': budget,
            'expenses': expenses, 'margin': margin,
            'margin_pct': round(margin_pct, 2),
        })
        total_budget += budget
        total_expenses += expenses

    overall_margin = total_budget - total_expenses
    overall_margin_pct = float(overall_margin / total_budget * 100) if total_budget > 0 else 0

    categories_breakdown = Costo.objects.values('categoria').annotate(
        total=Sum('monto'), count=Count('id')
    ).order_by('-total')

    context = {
        'summary_list': summary_list,
        'total_budget': total_budget,
        'total_expenses': total_expenses,
        'overall_margin': overall_margin,
        'overall_margin_pct': round(overall_margin_pct, 2),
        'categories_breakdown': categories_breakdown,
    }
    return render(request, 'core/cost_center_overview.html', context)


def _run_excel_import(control_source, fmr_source, clear_db=False):
    if clear_db:
        Entrega.objects.all().delete()
        FMR.objects.all().delete()
        Costo.objects.all().delete()
        OrdenCompra.objects.all().delete()

    def clean_val(val, default=""):
        if pd.isna(val) or val is None:
            return default
        if isinstance(val, str):
            return val.strip()
        return str(val)

    def clean_decimal(val):
        if pd.isna(val) or val is None:
            return None
        try:
            return float(str(val).replace("$", "").replace(" ", "").replace(",", ""))
        except ValueError:
            return None

    def clean_int(val):
        if pd.isna(val) or val is None:
            return None
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None

    def clean_date(val):
        if pd.isna(val) or val is None:
            return None
        if hasattr(val, 'date'):
            return val.date()
        try:
            s = str(val).strip()
            if not s or s.lower() == 'nan':
                return None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
        except Exception:
            pass
        return None

    # Load Control OC
    df_oc = pd.read_excel(control_source, sheet_name="Control OC")
    
    col_map = {}
    for c in df_oc.columns:
        c_clean = str(c).strip().upper().replace(' ', '').replace('\ufffd', '')
        if 'N' in c_clean and 'OC' in c_clean:
            col_map[c] = 'N_OC'
        elif 'CLIENTE' in c_clean:
            col_map[c] = 'CLIENTE'
        elif 'FECHA' in c_clean and 'OC' in c_clean:
            col_map[c] = 'FECHA_OC'
        elif 'PROYECTO' in c_clean:
            col_map[c] = 'PROYECTO'
        elif 'DESCRIP' in c_clean:
            col_map[c] = 'DESCRIPCION'
        elif 'VALOR' in c_clean:
            col_map[c] = 'VALOR_TOTAL'
        elif 'TIEMPO' in c_clean:
            col_map[c] = 'TIEMPO_FABRICACION'
        elif 'COMPROMISO' in c_clean:
            col_map[c] = 'FECHA_COMPROMISO'
        elif 'ESTADO' in c_clean:
            col_map[c] = 'ESTADO'
        elif 'PORCENTAJE' in c_clean or '%' in c_clean:
            col_map[c] = 'PORCENTAJE_ENTREGADO'
        elif 'ULTIMA' in c_clean or 'LTIMA' in c_clean:
            col_map[c] = 'FECHA_ULTIMA_ENTREGA'
        elif 'RESTANTES' in c_clean:
            col_map[c] = 'DIAS_RESTANTES'
        elif 'PRIORIDAD' in c_clean:
            col_map[c] = 'PRIORIDAD'
        elif 'GU' in c_clean and 'DESPACHO' in c_clean:
            col_map[c] = 'GUIA_DESPACHO_RESUMEN'
        elif 'FACTURA' in c_clean and 'FECHA' in c_clean:
            col_map[c] = 'FECHA_FACTURA'
        elif 'FACTURA' in c_clean:
            col_map[c] = 'FACTURA_RESUMEN'
        elif 'OBSERVACION' in c_clean:
            col_map[c] = 'OBSERVACIONES'
        elif c_clean == 'OC':
            col_map[c] = 'OC_LINK'
        elif c_clean == 'FMR':
            col_map[c] = 'FMR_LINK'
        elif c_clean == 'PLANO':
            col_map[c] = 'PLANO_LINK'
        elif c_clean == 'EXCEL':
            col_map[c] = 'EXCEL_LINK'
        elif c_clean == 'DOSSIER':
            col_map[c] = 'DOSSIER_LINK'

    df_oc.rename(columns=col_map, inplace=True)

    created_ocs = 0
    for _, row in df_oc.iterrows():
        oc_num = clean_val(row.get("N_OC"))
        if not oc_num or oc_num.lower() == 'nan':
            continue
        pct_val = row.get("PORCENTAJE_ENTREGADO")
        try:
            pct = float(pct_val)
            porcentaje = pct * 100 if pct <= 1.0 else pct
        except (TypeError, ValueError):
            porcentaje = 0.0

        oc_link = clean_val(row.get("OC_LINK"))
        fmr_link = clean_val(row.get("FMR_LINK"))
        plano_link = clean_val(row.get("PLANO_LINK"))
        excel_link = clean_val(row.get("EXCEL_LINK"))
        dossier_link = clean_val(row.get("DOSSIER_LINK"))

        OrdenCompra.objects.update_or_create(
            numero_oc=oc_num,
            defaults={
                'cliente': clean_val(row.get("CLIENTE"), default="Desconocido"),
                'fecha_oc': clean_date(row.get("FECHA_OC")),
                'proyecto': clean_val(row.get("PROYECTO")),
                'descripcion': clean_val(row.get("DESCRIPCION")),
                'valor_total': clean_decimal(row.get("VALOR_TOTAL")),
                'tiempo_fabricacion': clean_int(row.get("TIEMPO_FABRICACION")),
                'fecha_compromiso': clean_date(row.get("FECHA_COMPROMISO")),
                'estado': clean_val(row.get("ESTADO"), default="En proceso"),
                'porcentaje_entregado': porcentaje,
                'fecha_ultima_entrega': clean_date(row.get("FECHA_ULTIMA_ENTREGA")),
                'dias_restantes': clean_int(row.get("DIAS_RESTANTES")),
                'prioridad': clean_val(row.get("PRIORIDAD")),
                'guia_despacho_resumen': clean_val(row.get("GUIA_DESPACHO_RESUMEN")),
                'factura_resumen': clean_val(row.get("FACTURA_RESUMEN")),
                'fecha_factura': clean_date(row.get("FECHA_FACTURA")),
                'observaciones': clean_val(row.get("OBSERVACIONES")),
                'oc_link': oc_link if oc_link.startswith("http") else None,
                'fmr_link': fmr_link if fmr_link.startswith("http") else None,
                'plano_link': plano_link if plano_link.startswith("http") else None,
                'excel_link': excel_link if excel_link.startswith("http") else None,
                'dossier_link': dossier_link if dossier_link.startswith("http") else None,
            }
        )
        created_ocs += 1

    df_e = pd.read_excel(control_source, sheet_name="Entregas")
    
    col_map_entregas = {}
    for c in df_e.columns:
        c_clean = str(c).strip().upper().replace(' ', '').replace('\ufffd', '')
        if 'N' in c_clean and 'OC' in c_clean:
            col_map_entregas[c] = 'N_OC'
        elif 'FECHA' in c_clean:
            col_map_entregas[c] = 'FECHA_ENTREGA'
        elif 'GU' in c_clean:
            col_map_entregas[c] = 'GUIA_DESPACHO'
        elif 'CANTIDAD' in c_clean:
            col_map_entregas[c] = 'CANTIDAD_ENTREGADA'
        elif 'OBSERVACION' in c_clean:
            col_map_entregas[c] = 'OBSERVACIONES'
        elif 'ESTADO' in c_clean:
            col_map_entregas[c] = 'ESTADO'

    df_e.rename(columns=col_map_entregas, inplace=True)

    created_e = 0
    for _, row in df_e.iterrows():
        oc_num = clean_val(row.get("N_OC"))
        if not oc_num or oc_num.lower() == 'nan':
            continue
        oc, _ = OrdenCompra.objects.get_or_create(
            numero_oc=oc_num,
            defaults={'cliente': "FLUOR SALFA LTDA", 'descripcion': "Generado desde Entregas"}
        )
        raw_estado = clean_val(row.get("ESTADO")).upper()
        if 'INCOMPLETA' in raw_estado:
            estado = 'INCOMPLETA'
        elif 'FACTURADO' in raw_estado:
            estado = 'FACTURADO'
        else:
            estado = 'COMPLETA'
        Entrega.objects.create(
            orden_compra=oc,
            fecha_entrega=clean_date(row.get("FECHA_ENTREGA")),
            guia_despacho=clean_val(row.get("GUIA_DESPACHO")),
            cantidad_entregada=clean_val(row.get("CANTIDAD_ENTREGADA")),
            observaciones=clean_val(row.get("OBSERVACIONES")),
            estado=estado,
        )
        created_e += 1

    df_fmr = pd.read_excel(fmr_source, sheet_name="Hoja1", header=3)
    
    col_map_fmr = {}
    for c in df_fmr.columns:
        c_clean = str(c).strip().upper().replace(' ', '').replace('\ufffd', '')
        if c_clean == 'FMR':
            col_map_fmr[c] = 'FMR_CODE'
        elif 'FECHA' in c_clean:
            col_map_fmr[c] = 'FECHA'
        elif 'COTIZACION' in c_clean or 'COTIZACI' in c_clean:
            col_map_fmr[c] = 'COTIZACION'
        elif 'ORDEN' in c_clean or 'COMPRA' in c_clean or 'OC' in c_clean:
            col_map_fmr[c] = 'N_OC'
        elif 'GUIA' in c_clean or 'GUA' in c_clean:
            col_map_fmr[c] = 'GUIA_DESPACHO'
        elif 'FACTURA' in c_clean:
            col_map_fmr[c] = 'FACTURA'
        elif 'REGISTRO' in c_clean:
            col_map_fmr[c] = 'REGISTRO_LINK'

    df_fmr.rename(columns=col_map_fmr, inplace=True)

    created_fmrs = 0
    for _, row in df_fmr.iterrows():
        fmr_code = clean_val(row.get("FMR_CODE"))
        if not fmr_code or fmr_code.lower() == 'nan':
            continue
        oc_num = clean_val(row.get("N_OC"))
        oc = None
        if oc_num and oc_num.lower() != 'nan':
            oc, _ = OrdenCompra.objects.get_or_create(
                numero_oc=oc_num.strip(),
                defaults={'cliente': "FLUOR SALFA LTDA", 'descripcion': f"Generado desde FMR {fmr_code}"}
            )
        registro_link = clean_val(row.get("REGISTRO_LINK"))
        FMR.objects.update_or_create(
            fmr_code=fmr_code,
            defaults={
                'orden_compra': oc,
                'fecha': clean_date(row.get("FECHA")),
                'cotizacion': clean_val(row.get("COTIZACION")),
                'guia_despacho': clean_val(row.get("GUIA_DESPACHO")),
                'factura': clean_val(row.get("FACTURA")),
                'registro_link': registro_link if registro_link.startswith("http") else None,
            }
        )
        created_fmrs += 1

    # Recalculate progress
    for oc in OrdenCompra.objects.all():
        oc.recalcular_porcentaje()

    return {'ocs': created_ocs, 'deliveries': created_e, 'fmrs': created_fmrs}

@login_required
def import_data(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        clear_db = request.POST.get('clear_db') == 'true'

        if action == 'local':
            control_source = r"c:\Users\Coalfa\Desktop\FLUOR\Control_Ordenes_Compra_Maestranza.xlsx"
            fmr_source = r"c:\Users\Coalfa\Desktop\FLUOR\FLUOR FMR.xlsx"
            if not os.path.exists(control_source) or not os.path.exists(fmr_source):
                messages.error(request, "No se encontraron los archivos locales.")
                return render(request, 'core/import_data.html')
        elif action == 'upload':
            control_source = request.FILES.get('control_file')
            fmr_source = request.FILES.get('fmr_file')
            if not control_source or not fmr_source:
                messages.error(request, "Debe subir ambos archivos Excel.")
                return render(request, 'core/import_data.html')
        else:
            messages.error(request, "Acción no reconocida.")
            return render(request, 'core/import_data.html')

        try:
            results = _run_excel_import(control_source, fmr_source, clear_db=clear_db)
            messages.success(
                request,
                f"✅ Sincronización exitosa — {results['ocs']} OCs, "
                f"{results['deliveries']} Despachos, {results['fmrs']} FMRs importados."
            )
        except Exception as e:
            messages.error(request, f"Error durante la sincronización: {e}")

        return redirect('project_list')

    return render(request, 'core/import_data.html')


# ──────────────────────────────────────────────────────────────────────────────
# PACKING LIST
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def add_packing_item(request, numero_oc, entrega_id):
    orden_compra = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    entrega = get_object_or_404(Entrega, id=entrega_id, orden_compra=orden_compra)

    if request.method == 'POST':
        form = PackingListItemForm(request.POST, orden_compra=orden_compra)
        if form.is_valid():
            packing_item = form.save(commit=False)
            packing_item.entrega = entrega
            
            # --- AUTO-POPULAR CAMPOS DESDE ItemOC SI ESTÁN VACÍOS ---
            if packing_item.item_oc:
                if not packing_item.diametro and packing_item.item_oc.size_code:
                    packing_item.diametro = packing_item.item_oc.size_code
                if not packing_item.peso_kg and packing_item.item_oc.peso_unitario_kg:
                    packing_item.peso_kg = packing_item.item_oc.peso_unitario_kg
                if not packing_item.modelo_soporte and packing_item.item_oc.descripcion:
                    packing_item.modelo_soporte = packing_item.item_oc.descripcion
                if not packing_item.unidades and packing_item.item_oc.unidad:
                    packing_item.unidades = packing_item.item_oc.unidad
            # --------------------------------------------------------
            
            packing_item.save()  # triggers ItemOC.cantidad_entregada + OC % recalc

            registrar_trazabilidad(
                orden_compra=orden_compra,
                accion="Ítem Despachado (Packing List)",
                detalle=(
                    f"Ítem: '{packing_item.item_oc.descripcion}' — "
                    f"Cant: {packing_item.cantidad} | "
                    f"Bulto: {packing_item.numero_bulto or 'N/A'} | "
                    f"Guía: {entrega.guia_despacho or 'S/N'}"
                ),
                usuario=request.user
            )
            messages.success(request, "✅ Ítem añadido al Packing List del despacho.")
        else:
            messages.error(request, f"Error al añadir ítem: {form.errors}")

    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def delete_packing_item(request, numero_oc, item_id):
    orden_compra = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    packing_item = get_object_or_404(
        PackingListItem, id=item_id, entrega__orden_compra=orden_compra
    )

    registrar_trazabilidad(
        orden_compra=orden_compra,
        accion="Eliminar Ítem Packing List",
        detalle=(
            f"Se eliminó '{packing_item.item_oc.descripcion}' del despacho "
            f"guía {packing_item.entrega.guia_despacho or 'S/N'}. "
            f"Cantidad: {packing_item.cantidad}."
        ),
        usuario=request.user
    )
    packing_item.delete()  # triggers ItemOC.cantidad_entregada + OC % recalc
    messages.success(request, "🗑️ Ítem de Packing List eliminado.")
    return redirect('project_detail', numero_oc=numero_oc)


# ──────────────────────────────────────────────────────────────────────────────
# DETAILED COSTS ACTIONS (POST only)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def add_cost_material(request, numero_oc):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        form = MateriaPrimaForm(request.POST)
        if form.is_valid():
            mat = form.save(commit=False)
            mat.orden_compra = project
            if not mat.cantidad:
                # Si no se ingresa cantidad, usar total fijado directamente
                mat.total = form.cleaned_data.get('total') or mat.valor_unitario
            mat.save()
            
            registrar_trazabilidad(
                project,
                "Compra de Material",
                f"Se compró materia prima '{mat.producto}' (Cant: {mat.cantidad or '—'} | Unit: ${mat.valor_unitario:,.0f} | Total: ${mat.total:,.0f}).",
                request.user
            )
            messages.success(request, f"✅ Material '{mat.producto}' registrado exitosamente.")
        else:
            messages.error(request, f"Error al registrar material: {form.errors}")
    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def delete_cost_material(request, numero_oc, item_id):
    project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    mat = get_object_or_404(MateriaPrima, id=item_id, orden_compra=project)
    
    registrar_trazabilidad(
        project,
        "Eliminación de Material",
        f"Se eliminó registro de materia prima '{mat.producto}' por ${mat.total:,.0f}.",
        request.user
    )
    mat.delete()
    messages.success(request, "🗑️ Registro de material eliminado.")
    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def add_cost_mano_obra(request, numero_oc):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        form = ManoDeObraForm(request.POST)
        if form.is_valid():
            mo = form.save(commit=False)
            mo.orden_compra = project
            mo.save()
            
            registrar_trazabilidad(
                project,
                "Costo de Mano de Obra",
                f"Se registró mano de obra para '{mo.cargo.nombre}' (Días: {mo.dias} | Horas/d: {mo.horas} | Trab: {mo.cantidad_trabajadores} | Total: ${mo.total:,.0f}).",
                request.user
            )
            messages.success(request, f"✅ Mano de Obra para '{mo.cargo.nombre}' registrada exitosamente.")
        else:
            messages.error(request, f"Error al registrar mano de obra: {form.errors}")
    return redirect('project_detail', numero_oc=numero_oc)


@login_required
def delete_cost_mano_obra(request, numero_oc, item_id):
    project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    mo = get_object_or_404(ManoDeObra, id=item_id, orden_compra=project)
    
    registrar_trazabilidad(
        project,
        "Eliminación de Mano de Obra",
        f"Se eliminó registro de mano de obra de '{mo.cargo.nombre}' por ${mo.total:,.0f}.",
        request.user
    )
    mo.delete()
    messages.success(request, "🗑️ Registro de mano de obra eliminado.")
    return redirect('project_detail', numero_oc=numero_oc)


# VISTAS DE COSTEO DETALLADO / RENDICIÓN Y PACKING LISTS

@login_required
def project_rendicion(request, numero_oc):
    project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    manos_obra = project.manos_de_obra.select_related('cargo').all()
    materias_primas = project.materias_primas.all()
    cargos = Cargo.objects.all()
    
    context = {
        'project': project,
        'manos_obra': manos_obra,
        'materias_primas': materias_primas,
        'cargos': cargos,
    }
    return render(request, 'core/project_rendicion.html', context)


@login_required
def add_materia_prima(request, numero_oc):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        producto = request.POST.get('producto')
        cantidad = request.POST.get('cantidad')
        valor_unitario = request.POST.get('valor_unitario') or 0
        total_p = request.POST.get('total')
        
        qty = Decimal(cantidad) if (cantidad and cantidad.strip()) else None
        val = Decimal(valor_unitario)
        
        mp = MateriaPrima(
            orden_compra=project,
            producto=producto,
            cantidad=qty,
            valor_unitario=val
        )
        if qty is None:
            mp.total = Decimal(total_p) if (total_p and total_p.strip()) else val
        mp.save()
        
        registrar_trazabilidad(
            project,
            "Carga de Materia Prima (Rendición)",
            f"Se cargó materia prima '{mp.producto}' por un total de ${mp.total:,.0f}.",
            request.user
        )
        messages.success(request, f"✅ Materia prima '{mp.producto}' registrada con éxito.")
    return redirect('project_rendicion', numero_oc=numero_oc)


@login_required
def delete_materia_prima(request, numero_oc, item_id):
    project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    mp = get_object_or_404(MateriaPrima, id=item_id, orden_compra=project)
    
    registrar_trazabilidad(
        project,
        "Eliminación Materia Prima (Rendición)",
        f"Se eliminó '{mp.producto}' con costo total ${mp.total:,.0f}.",
        request.user
    )
    mp.delete()
    messages.success(request, "🗑️ Registro de materia prima eliminado.")
    return redirect('project_rendicion', numero_oc=numero_oc)


@login_required
def add_mano_obra_detallada(request, numero_oc):
    if request.method == 'POST':
        project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
        cargo_id = request.POST.get('cargo')
        cargo_obj = get_object_or_404(Cargo, id=cargo_id)
        dias = int(request.POST.get('dias') or 1)
        horas = int(request.POST.get('horas') or 8)
        cantidad_trabajadores = int(request.POST.get('cantidad_trabajadores') or 1)
        horas_extra = int(request.POST.get('horas_extra') or 0)
        
        mo = ManoDeObra(
            orden_compra=project,
            cargo=cargo_obj,
            dias=dias,
            horas=horas,
            cantidad_trabajadores=cantidad_trabajadores,
            horas_extra=horas_extra
        )
        mo.save()
        
        registrar_trazabilidad(
            project,
            "Carga Mano de Obra (Rendición)",
            f"Se cargó '{mo.cargo.nombre}' por total ${mo.total:,.0f}.",
            request.user
        )
        messages.success(request, f"✅ Mano de obra para '{mo.cargo.nombre}' registrada.")
    return redirect('project_rendicion', numero_oc=numero_oc)


@login_required
def delete_mano_obra_detallada(request, numero_oc, item_id):
    project = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    mo = get_object_or_404(ManoDeObra, id=item_id, orden_compra=project)
    
    registrar_trazabilidad(
        project,
        "Eliminación Mano de Obra (Rendición)",
        f"Se eliminó '{mo.cargo.nombre}' por total ${mo.total:,.0f}.",
        request.user
    )
    mo.delete()
    messages.success(request, "🗑️ Registro de mano de obra detallada eliminado.")
    return redirect('project_rendicion', numero_oc=numero_oc)


# WEASYPRINT / REPORTLAB PACKING LIST

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from django.http import HttpResponse

@login_required
def generate_packing_list_pdf(request, packing_list_id):
    pl = get_object_or_404(PackingList, id=packing_list_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="packing_list_{pl.numero_correlativo}.pdf"'
    
    # Crear documento PDF
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=18,
        textColor=colors.HexColor('#0d1220'),
        alignment=1 # Centro
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )
    
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=10.5,
        textColor=colors.HexColor('#333333')
    )

    # 1. Membrete e Información
    empresa_info = f"<b>{pl.empresa}</b><br/>Giro: Maestranza y Fabricaciones Metálicas<br/>Dir: {pl.direccion}<br/>Correo: {pl.correo}<br/>Fono: {pl.telefono}"
    documento_info = f"<font color='#0D1220'>Packing List N° {pl.numero_correlativo:05d}</font><br/><br/><b>CONTROL DE CALIDAD<br/>PACKING LIST</b>"
    
    empresa_data = [
        [
            Paragraph(empresa_info, body_style),
            Paragraph(documento_info, title_style)
        ]
    ]
    empresa_table = Table(empresa_data, colWidths=[250, 280])
    empresa_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(empresa_table)
    story.append(Spacer(1, 10))
    
    # 2. Información del Cliente
    cliente_data = [
        [
            Paragraph(f"<b>Cliente:</b> {pl.nombre_cliente}", body_style),
            Paragraph(f"<b>Fecha Orden:</b> {pl.fecha_orden.strftime('%d-%m-%Y') if pl.fecha_orden else 'N/A'}", body_style)
        ],
        [
            Paragraph(f"<b>N° OC Asociada:</b> {pl.orden_compra.numero_oc}", body_style),
            Paragraph(f"<b>Fecha Envío:</b> {pl.fecha_envio.strftime('%d-%m-%Y') if pl.fecha_envio else 'N/A'}", body_style)
        ]
    ]
    if pl.entrega:
        cliente_data.append([
            Paragraph(f"<b>Guía Despacho:</b> {pl.entrega.guia_despacho or 'N/A'}", body_style),
            Paragraph("", body_style)
        ])
        
    cliente_table = Table(cliente_data, colWidths=[265, 265])
    cliente_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#dddddd')),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f9f9f9')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(cliente_table)
    story.append(Spacer(1, 15))
    
    # 3. Datos de la tabla de Ítems — cabeceras según tipo_medida
    col_m1 = pl.col_medida_1  # 'Ø' o 'L'
    col_m2 = pl.col_medida_2  # 'ALTO' o 'H'
    table_data = [[
        Paragraph('Ítem / Descripción', header_style),
        Paragraph('Modelo Soporte', header_style),
        Paragraph(col_m1, header_style),
        Paragraph(col_m2, header_style),
        Paragraph('Estado', header_style),
        Paragraph('Unidades', header_style)
    ]]
    
    items = []
    if pl.entrega:
         items = pl.entrega.packing_list_items.select_related('item_oc').all()
         
    for index, item in enumerate(items, start=1):
        desc = item.item_oc.descripcion
        # Prefer numeric medida_1/medida_2; fall back to legacy text fields
        m1 = str(item.medida_1) if item.medida_1 is not None else (item.diametro or 'N/A')
        m2 = str(item.medida_2) if item.medida_2 is not None else (item.alto_item or 'N/A')
        table_data.append([
            Paragraph(f"{index}. {desc}", body_style),
            Paragraph(item.modelo_soporte or "N/A", body_style),
            Paragraph(m1, body_style),
            Paragraph(m2, body_style),
            Paragraph(item.estado_item or "N/A", body_style),
            Paragraph(item.unidades or str(int(item.cantidad)), body_style)
        ])
        
    if not items:
        table_data.append([Paragraph("No se encontraron ítems en esta entrega.", body_style), "", "", "", "", ""])
        
    items_table = Table(table_data, colWidths=[180, 110, 60, 60, 60, 60])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d1220')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9f9')]),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 35))
    
    # 4. Firma
    firma_data = [
        ["", ""],
        ["_________________________", ""],
        ["JEFE DE OPERACIONES", ""],
        ["MAESTRANZA BARK SPA", ""]
    ]
    firma_table = Table(firma_data, colWidths=[265, 265])
    firma_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,2), (0,2), 'Helvetica-Bold'),
        ('FONTSIZE', (0,2), (0,2), 8.5),
    ]))
    story.append(KeepTogether([firma_table]))
    
    doc.build(story)
    return response


@login_required
def create_packing_list(request, numero_oc, entrega_id):
    orden_compra = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    entrega = get_object_or_404(Entrega, id=entrega_id, orden_compra=orden_compra)
    
    if request.method == 'POST':
        fecha_orden = request.POST.get('fecha_orden') or timezone.now().date()
        fecha_envio = request.POST.get('fecha_envio') or timezone.now().date()
        nombre_cliente = request.POST.get('nombre_cliente') or orden_compra.cliente
        
        pl = PackingList(
            orden_compra=orden_compra,
            entrega=entrega,
            fecha_orden=fecha_orden,
            fecha_envio=fecha_envio,
            nombre_cliente=nombre_cliente
        )
        pl.save()
        
        registrar_trazabilidad(
            orden_compra,
            "Packing List Creado",
            f"Se generó Packing List N° {pl.numero_correlativo} para despacho guía {entrega.guia_despacho or 'S/N'}.",
            request.user
        )
        messages.success(request, f"✅ Packing List N° {pl.numero_correlativo} creado exitosamente.")
        
    return redirect('entrega_detail', numero_oc=numero_oc, entrega_id=entrega_id)


@login_required
def entrega_detail(request, numero_oc, entrega_id):
    orden_compra = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    entrega = get_object_or_404(Entrega, id=entrega_id, orden_compra=orden_compra)
    items = entrega.packing_list_items.select_related('item_oc').all()
    packing_lists = entrega.packing_lists.all()
    
    if request.method == 'POST' and 'cambiar_facturacion' in request.POST:
        nuevo_res = request.POST.get('estado_facturacion')
        if nuevo_res in dict(Entrega.ESTADOS_FACTURACION):
            entrega.estado_facturacion = nuevo_res
            entrega.save()
            registrar_trazabilidad(
                orden_compra,
                "Control Facturación Guía",
                f"Guía {entrega.guia_despacho}: estado facturación ahora es '{entrega.get_estado_facturacion_display()}'.",
                request.user
            )
            messages.success(request, f"✅ Estado de facturación de despacho cambiado a '{entrega.get_estado_facturacion_display()}'.")
            return redirect('entrega_detail', numero_oc=numero_oc, entrega_id=entrega_id)

    # Si se añade un packing_list_item directamente desde la vista de detalle
    if request.method == 'POST' and 'agregar_item' in request.POST:
        form = PackingListItemForm(request.POST, orden_compra=orden_compra)
        if form.is_valid():
            p_item = form.save(commit=False)
            p_item.entrega = entrega
            p_item.save()
            registrar_trazabilidad(
                orden_compra,
                "Ítem Añadido a Despacho",
                f"U: {p_item.cantidad} de '{p_item.item_oc.descripcion}' agregados a guía {entrega.guia_despacho}.",
                request.user
            )
            messages.success(request, "✅ Ítem agregado al despacho.")
            return redirect('entrega_detail', numero_oc=numero_oc, entrega_id=entrega_id)
        else:
            messages.error(request, f"Error al agregar ítem: {form.errors}")

    form = PackingListItemForm(orden_compra=orden_compra)
    context = {
        'orden_compra': orden_compra,
        'entrega': entrega,
        'items': items,
        'packing_lists': packing_lists,
        'estados_facturacion': Entrega.ESTADOS_FACTURACION,
        'form': form,
    }
    return render(request, 'core/entrega_detail.html', context)


@login_required
def despachos_list(request):
    filtro = request.GET.get('filtro', 'por_facturar')
    query = request.GET.get('q', '')

    entregas = Entrega.objects.select_related('orden_compra').prefetch_related('packing_list_items__item_oc', 'facturas').all().order_by('-fecha_entrega')

    if query:
        entregas = entregas.filter(
            Q(guia_despacho__icontains=query) |
            Q(orden_compra__numero_oc__icontains=query) |
            Q(orden_compra__cliente__icontains=query)
        )

    if filtro == 'por_facturar':
        entregas = entregas.filter(estado_facturacion='por_facturar')
    elif filtro == 'facturado':
        entregas = entregas.filter(estado_facturacion='facturado')

    if request.method == 'POST' and 'facturar_entrega' in request.POST:
        entrega_id = request.POST.get('entrega_id')
        num_factura = request.POST.get('numero_factura')
        monto = request.POST.get('monto')
        fecha_emision = request.POST.get('fecha_emision') or timezone.now().date()

        entrega = get_object_or_404(Entrega, id=entrega_id)

        if num_factura and monto:
            # Crear Factura
            Factura.objects.create(
                orden_compra=entrega.orden_compra,
                entrega=entrega,
                numero_factura=num_factura,
                monto=Decimal(monto),
                fecha_emision=fecha_emision,
                estado='pendiente'
            )
            entrega.estado_facturacion = 'facturado'
            entrega.save()

            # Registrar trazabilidad
            registrar_trazabilidad(
                entrega.orden_compra,
                "Guía Facturada",
                f"Se asoció la Factura N° {num_factura} por ${float(monto):,.0f} al despacho guía {entrega.guia_despacho} y se marcó como Facturado.",
                request.user
            )
            messages.success(request, f"✅ Despacho GD-{entrega.guia_despacho or 'S/N'} facturado exitosamente con Fac N° {num_factura}.")
        else:
            entrega.estado_facturacion = 'facturado'
            entrega.save()
            messages.success(request, f"✅ Despacho GD-{entrega.guia_despacho or 'S/N'} marcado como Facturado.")

        return redirect(f"{request.path}?filtro={filtro}&q={query}")

    if request.method == 'POST' and 'desfacturar_entrega' in request.POST:
        entrega_id = request.POST.get('entrega_id')
        entrega = get_object_or_404(Entrega, id=entrega_id)
        entrega.estado_facturacion = 'por_facturar'
        entrega.save()

        # Opcionalmente cancelar facturas asociadas a esta entrega para mantener costos consistentes
        for f in entrega.facturas.all():
            f.delete()

        registrar_trazabilidad(
            entrega.orden_compra,
            "Despacho por Facturar",
            f"Se cambió estado de de despacho guía {entrega.guia_despacho} a 'Por facturar' (se eliminaron facturas asociadas).",
            request.user
        )
        messages.success(request, f"🔄 Despacho GD-{entrega.guia_despacho or 'S/N'} devuelto a 'Por Facturar'.")
        return redirect(f"{request.path}?filtro={filtro}&q={query}")

    context = {
        'entregas': entregas,
        'filtro': filtro,
        'q': query,
    }
    return render(request, 'core/despachos_list.html', context)


# ──────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN EXCEL PACKING LIST (Formato imagen del cliente)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def export_packing_list_excel(request, packing_list_id):
    from openpyxl.styles import Border, Side
    pl = get_object_or_404(PackingList, id=packing_list_id)
    oc = pl.orden_compra

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"PL-{pl.numero_correlativo:05d}"

    # ── Estilos base ──────────────────────────────────────────────────────────
    thin = Side(style='thin', color='AAAAAA')
    thick = Side(style='medium', color='1a3a5c')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    def cell(ws, row, col, value='', bold=False, size=10, color='000000',
             bg=None, align='left', border=None, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, color=color, name='Calibri')
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if bg:
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        if border:
            c.border = border
        return c

    # ── Anchos de columna (A-H) ───────────────────────────────────────────────
    col_widths = {'A': 8, 'B': 22, 'C': 10, 'D': 14, 'E': 16, 'F': 14, 'G': 14, 'H': 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── FILA 1: Header "CONTROL DE CALIDAD" ───────────────────────────────────
    ws.merge_cells('D1:H1')
    cell(ws, 1, 4, 'CONTROL DE CALIDAD', bold=True, size=11, color='1a3a5c', align='center', bg='D6E4F0')
    ws.merge_cells('A1:C4')  # Logo area
    logo_cell = ws.cell(row=1, column=1)
    logo_cell.value = 'MAESTRANZA\nBARK SPA'
    logo_cell.font = Font(bold=True, size=16, name='Calibri', color='1a3a5c')
    logo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    logo_cell.fill = PatternFill(start_color='EAF0FA', end_color='EAF0FA', fill_type='solid')

    # ── FILA 2-4: "PACKING LIST" grande ──────────────────────────────────────
    ws.merge_cells('D2:H4')
    pl_cell = ws.cell(row=2, column=4)
    pl_cell.value = 'PACKING LIST'
    pl_cell.font = Font(bold=True, size=22, name='Calibri', color='1a3a5c')
    pl_cell.alignment = Alignment(horizontal='center', vertical='center')
    pl_cell.fill = PatternFill(start_color='EAF0FA', end_color='EAF0FA', fill_type='solid')

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22

    # ── FILA 5: Separador ─────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 8

    # ── FILAS 6-10: Datos empresa y cliente ───────────────────────────────────
    empresa_data = [
        (f'Empresa: {pl.empresa}', f'Nombre de cliente: {pl.nombre_cliente}'),
        (f'Dirección: {pl.direccion}',   f'Fecha de orden: {pl.fecha_orden.strftime("%d/%m/%Y") if pl.fecha_orden else "N/A"}'),
        (f'Ciudad: Calama',               f'Fecha de envío: {pl.fecha_envio.strftime("%d/%m/%Y") if pl.fecha_envio else "N/A"}'),
        (f'Correo: {pl.correo}',          ''),
        (f'Teléfono: {pl.telefono}',      ''),
    ]

    for i, (izq, der) in enumerate(empresa_data, start=6):
        ws.merge_cells(f'A{i}:C{i}')
        c_izq = ws.cell(row=i, column=1, value=izq)
        c_izq.font = Font(size=9, name='Calibri')
        c_izq.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'D{i}:H{i}')
        c_der = ws.cell(row=i, column=4, value=der)
        c_der.font = Font(size=9, name='Calibri', bold=True if der.startswith('Nombre') or der.startswith('Fecha') else False)
        c_der.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[i].height = 15

    # ── FILA 11: Separador ────────────────────────────────────────────────────
    ws.row_dimensions[11].height = 8

    # ── FILAS 12-14: Orden de Compra ─────────────────────────────────────────
    ws.merge_cells('A12:C12')
    c_oc_hdr = ws.cell(row=12, column=1, value='Orden de compra')
    c_oc_hdr.font = Font(bold=True, size=9, name='Calibri')
    c_oc_hdr.alignment = Alignment(horizontal='center', vertical='center')
    c_oc_hdr.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    c_oc_hdr.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[12].height = 15

    ws.merge_cells('A13:C14')
    c_oc_val = ws.cell(row=13, column=1, value=oc.numero_oc)
    c_oc_val.font = Font(size=9, name='Calibri')
    c_oc_val.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_oc_val.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[13].height = 22
    ws.row_dimensions[14].height = 22

    # ── FILA 15: Separador ────────────────────────────────────────────────────
    ws.row_dimensions[15].height = 10

    # ── FILA 16: Encabezados de tabla ─────────────────────────────────────────
    headers = ['ITEM', 'MODELO SOPORTE', 'Ø', 'ALTO', 'ESTADO', 'UNIDADES']
    col_map = [1, 2, 3, 4, 5, 6]
    hdr_bg = '1a3a5c'
    for i, h in enumerate(headers):
        c = ws.cell(row=16, column=col_map[i], value=h)
        c.font = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
        c.fill = PatternFill(start_color=hdr_bg, end_color=hdr_bg, fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border_thin
    ws.row_dimensions[16].height = 18

    # ── FILAS 17+: Ítems ──────────────────────────────────────────────────────
    items = []
    if pl.entrega:
        items = list(pl.entrega.packing_list_items.select_related('item_oc').all())

    row_num = 17
    alt_colors = ['FFFFFF', 'F2F7FB']
    if items:
        for idx, item in enumerate(items, start=1):
            bg = alt_colors[idx % 2]
            row_vals = [
                idx,
                item.modelo_soporte or item.item_oc.descripcion,
                item.diametro or '—',
                item.alto_item or '—',
                (item.estado_item or 'ENTREGADO').upper(),
                item.unidades or str(int(float(item.cantidad or 1)))
            ]
            for ci, val in enumerate(row_vals):
                c = ws.cell(row=row_num, column=col_map[ci], value=val)
                c.font = Font(size=9, name='Calibri')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                c.border = border_thin
            ws.row_dimensions[row_num].height = 16
            row_num += 1
    else:
        ws.merge_cells(f'A{row_num}:F{row_num}')
        c = ws.cell(row=row_num, column=1, value='Sin ítems registrados en este despacho')
        c.font = Font(italic=True, size=9, color='888888', name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1

    # ── Firma final ───────────────────────────────────────────────────────────
    firma_row = row_num + 3
    ws.merge_cells(f'D{firma_row}:F{firma_row}')
    ws.cell(row=firma_row, column=4, value='_________________________________') \
      .alignment = Alignment(horizontal='center')
    ws.row_dimensions[firma_row].height = 15

    ws.merge_cells(f'D{firma_row+1}:F{firma_row+1}')
    c_empresa = ws.cell(row=firma_row+1, column=4, value='MAESTRANZA BARK SPA')
    c_empresa.font = Font(bold=True, size=9, name='Calibri')
    c_empresa.alignment = Alignment(horizontal='center')

    ws.merge_cells(f'D{firma_row+2}:F{firma_row+2}')
    c_cargo = ws.cell(row=firma_row+2, column=4, value='JEFE DE OPERACIONES.')
    c_cargo.font = Font(size=9, name='Calibri')
    c_cargo.alignment = Alignment(horizontal='center')

    ws.merge_cells(f'D{firma_row+3}:F{firma_row+3}')
    ws.cell(row=firma_row+3, column=4, value='- ' * 20) \
      .alignment = Alignment(horizontal='center')

    # ── Respuesta ──────────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="PackingList_{pl.numero_correlativo:05d}.xlsx"'
    with io.BytesIO() as b:
        wb.save(b)
        response.write(b.getvalue())
    return response


# ──────────────────────────────────────────────────────────────────────────────
# EXPORTACIONES A EXCEL (CSV UTF-8)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def export_bom_csv(request, numero_oc):
    oc = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    # Excel no permite los caracteres \ / ? * [ ] en los nombres de hoja
    safe_title = "".join(c if c not in r'\/?*[]' else '-' for c in f"BOM {oc.numero_oc}")
    ws.title = safe_title[:31]  
    
    # Estilos Steel & Amber
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    
    headers = [
        'Línea', 'Cód. Plano / Alt', 'Item Code', 'Size Code', 
        'Marca / Descripción', 'UOM', 'Cant. Solicitada', 
        'Cant. Entregada', 'Peso Unitario (kg)', 'Peso Total (kg)'
    ]
    
    ws.append(headers)
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18

    # Ampliar la columna de descripción
    ws.column_dimensions['E'].width = 45
    
    for item in oc.items.all():
        ws.append([
            item.linea or '', 
            item.codigo or '', 
            item.item_code or '', 
            item.size_code or '', 
            item.descripcion or '', 
            item.uom or item.unidad or 'EA',
            float(item.cantidad or 0), 
            float(item.cantidad_entregada or 0),
            float(item.peso_unitario_kg or 0), 
            float(item.peso_total_kg or 0)
        ])
        
    # Formatear números
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="BOM_OC_{oc.numero_oc}.xlsx"'
    
    with io.BytesIO() as b:
        wb.save(b)
        response.write(b.getvalue())
        
    return response

@login_required
def export_rendicion_csv(request, numero_oc):
    oc = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendición de Costos"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid") # Amber
    dark_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    title_font = Font(bold=True, size=14)
    normal_bold = Font(bold=True)
    
    ws['A1'] = 'REPORTE DE RENDICIÓN DE COSTOS'
    ws['A1'].font = title_font
    ws['A2'] = f'OC: {oc.numero_oc}'
    ws['A3'] = f'CLIENTE: {oc.cliente}'
    if oc.proyecto:
        ws['A4'] = f'PROYECTO: {oc.proyecto}'
        
    current_row = 6
    
    def add_section_header(title, row):
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = header_font
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center")
        return row + 1

    current_row = add_section_header('RESUMEN FINANCIERO GLOBAL', current_row)
    
    fin_data = [
        ('Valor Total Adjudicado (Sin IVA):', float(oc.valor_total or 0)),
        ('Total Gastos en Materia Prima:', float(oc.costo_total_materiales or 0)),
        ('Total Gastos en Mano de Obra:', float(oc.costo_total_mano_obra or 0)),
        ('Costo Total del Trabajo:', float(oc.costo_total_trabajo or 0)),
        ('UTILIDAD REAL:', float(oc.utilidad_real or 0)),
        ('MARGEN DE UTILIDAD (%):', float(oc.porcentaje_utilidad or 0) / 100 if oc.porcentaje_utilidad else 0),
        ('Costo por Kilo:', float(oc.costo_por_kilo or 0))
    ]
    
    for label, val in fin_data:
        ws.cell(row=current_row, column=1, value=label).font = normal_bold
        val_cell = ws.cell(row=current_row, column=2, value=val)
        if 'RENDIM' in label or 'UTILIDAD (%)' in label:
            val_cell.number_format = '0.00%'
        else:
            val_cell.number_format = '#,##0'
        current_row += 1
        
    current_row += 1
    
    # ── Materias Primas
    current_row = add_section_header('DETALLE DE MATERIAS PRIMAS', current_row)
    headers_mp = ['Producto / Material', 'Cantidad', 'Valor Unitario ($)', 'Valor Total ($)']
    for col, h in enumerate(headers_mp, 1):
        c = ws.cell(row=current_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    
    current_row += 1
    for mp in oc.materias_primas.all():
        ws.cell(row=current_row, column=1, value=mp.producto)
        ws.cell(row=current_row, column=2, value=float(mp.cantidad or 1))
        ws.cell(row=current_row, column=3, value=float(mp.valor_unitario or 0)).number_format = '#,##0'
        ws.cell(row=current_row, column=4, value=float(mp.total or 0)).number_format = '#,##0'
        current_row += 1
        
    current_row += 2
    
    # ── Mano de Obra
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'DETALLE DE MANO DE OBRA'
    cell.font = header_font
    cell.fill = dark_fill
    cell.alignment = Alignment(horizontal="center")
    current_row += 1
    
    headers_mo = ['Cargo / Especialidad', 'Días Trabajados', 'Hrs Normales', 'Hrs Extra', 'Cant. Trabajadores', 'Costo Horario ($)', 'Total ($)']
    for col, h in enumerate(headers_mo, 1):
        c = ws.cell(row=current_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        
    current_row += 1
    for mo in oc.manos_de_obra.all():
        cargo_nombre = mo.cargo.nombre if mo.cargo else (mo.cargo_otro or 'N/A')
        costo_hr = mo.precio_hora or (mo.cargo.precio_por_hora if mo.cargo else 0)
        
        ws.cell(row=current_row, column=1, value=cargo_nombre)
        ws.cell(row=current_row, column=2, value=float(mo.dias or 0))
        ws.cell(row=current_row, column=3, value=float(mo.horas or 0))
        ws.cell(row=current_row, column=4, value=float(mo.horas_extra or 0))
        ws.cell(row=current_row, column=5, value=float(mo.cantidad_trabajadores or 1))
        ws.cell(row=current_row, column=6, value=float(costo_hr)).number_format = '#,##0'
        ws.cell(row=current_row, column=7, value=float(mo.total or 0)).number_format = '#,##0'
        current_row += 1
        
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['G'].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Rendicion_Costos_{oc.numero_oc}.xlsx"'
    
    with io.BytesIO() as b:
        wb.save(b)
        response.write(b.getvalue())
        
    return response


# ──────────────────────────────────────────────────────────────────────────────
# LECTURA AUTOMÁTICA DE DOCUMENTOS (API)
# ──────────────────────────────────────────────────────────────────────────────









@csrf_exempt
def api_analizar_documento(request):
    if request.method == 'POST' and request.FILES.get('documento'):
        archivo = request.FILES['documento']
        nombre = archivo.name.lower()
        
        datos = {
            'numero_oc': '',
            'cliente': '',
            'proyecto': '',
            'fecha_oc': '',
            'valor_total': '',
        }
        
        try:
            texto = ""
            texto_items = ""
            if nombre.endswith('.pdf'):
                with pdfplumber.open(archivo) as pdf:
                    for i, pagina in enumerate(pdf.pages):
                        extracted = pagina.extract_text()
                        if extracted:
                            texto += extracted + "\n"
                            # Guardar específicamente las últimas 4 páginas para los ítems
                            if len(pdf.pages) - i <= 4:
                                texto_items += extracted + "\n"
            elif nombre.endswith('.xlsx'):
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    linea = " ".join([str(v) for v in row if v is not None])
                    texto += linea + "\n"
                texto_items = texto
                    
            # Basic RegEx Extraction for OC Fields
            # Número OC (ej: D3MC104397, o PO-12345)
            # Many times it's near "Purchase Order", "OC", "Orden de Compra"
            re_oc = re.search(r'(?:OC|Orden de Compra|PO|Purchase Order)[\s:N°#]+([A-Z0-9\-\/]+)', texto, re.IGNORECASE)
            if re_oc:
                datos['numero_oc'] = re_oc.group(1).strip()
            
            # Cliente
            re_cliente = re.search(r'(?:Cliente|Customer|To)[\s:]+([A-Za-z0-9\.\s&]+)', texto, re.IGNORECASE)
            # We take just the first line
            if re_cliente:
                cliente_candidate = re_cliente.group(1).split('\n')[0].strip()
                if len(cliente_candidate) > 2:
                    datos['cliente'] = cliente_candidate[:100]
            
            # Proyecto
            re_proyecto = re.search(r'(?:Proyecto|Project)[\s:]+([^\n]+)', texto, re.IGNORECASE)
            if re_proyecto:
                datos['proyecto'] = re_proyecto.group(1).strip()[:100]
                
            # Fecha OC
            # Buscando formatos comunes de fecha dd/mm/yyyy o yyyy-mm-dd
            re_fecha = re.search(r'(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})', texto)
            if re_fecha:
                raw_date = re_fecha.group(1)
                datos['fecha_oc'] = raw_date
                
            # Monto total aproximado
            re_monto = re.search(r'(?:Total|Neto)[\s:\$]*([\d\.,]+)', texto, re.IGNORECASE)
            if re_monto:
                # remove non digits except dot or comma
                monto_val = re.sub(r'[^\d]', '', re_monto.group(1))
                if monto_val:
                    datos['valor_total'] = monto_val

            # Extraer ítems de material (BOM y FMR)
            items_extraidos = []
            fmr_mode = False
            current_fmr_item = None
            
            lineas = texto.split('\n')
            for i, linea in enumerate(lineas):
                linea_clean = linea.strip()
                if not linea_clean:
                    continue
                
                # --- FORMATO BARK OC (`Item code: ...`) ---
                if 'Item code:' in linea_clean:
                    match = re.search(r'Item code:\s+(\S+)\s+(\d{1,4})\s+(.+?)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([A-Za-z0-9\-]+)', linea_clean)
                    if match:
                        item_code = match.group(1)
                        linea_num = match.group(2).lstrip('0') or '1'
                        size = match.group(3).strip()
                        cant = match.group(4).replace(',', '')
                        precio_u = match.group(5).replace(',', '')
                        precio_t = match.group(6).replace(',', '')
                        fecha_e_raw = match.group(7).strip()
                        
                        desc = ""
                        uom = "EA"
                        
                        if i + 1 < len(lineas):
                            next_line = lineas[i+1].strip()
                            # Match something like "SOPORTE    EA - each    CLP"
                            match_next = re.search(r'^(.*?)\s+([A-Z]{2,4}(?:\s*-\s*[a-zA-Z]+)?)\s+([A-Z]{3,4})$', next_line)
                            if match_next:
                                desc = match_next.group(1).strip()
                                uom = match_next.group(2).strip()
                            elif next_line and not "Item code:" in next_line:
                                desc = next_line[:80] # Fallback a usar toda la línea como desc
                                
                        items_extraidos.append({
                            'linea': linea_num,
                            'item_code': item_code,
                            'size_code': size if '<No Size>' not in size else '',
                            'descripcion': desc[:255] if desc else f"Item {item_code}",
                            'cantidad': cant,
                            'uom': uom,
                            'precio_unitario': precio_u,
                            'precio_total': precio_t,
                            'fecha_entrega': fecha_e_raw
                        })
                    continue

                # --- FORMATO FMR (COMMODITY CODE) ---
                match_fmr = re.match(r'^([A-Z0-9\-]{5,})\s+(\d+)\s+([\d\.,]+)\s+([A-Za-z]{2,5}\.?)\s+(.*)$', linea_clean)
                if match_fmr and len(match_fmr.group(1)) > 5:
                    if current_fmr_item:
                        items_extraidos.append(current_fmr_item)
                    
                    fmr_mode = True
                    current_fmr_item = {
                        'item_code': match_fmr.group(1),
                        'codigo': match_fmr.group(1),
                        'linea': match_fmr.group(2),
                        'cantidad': match_fmr.group(3).replace(',', ''),
                        'uom': match_fmr.group(4),
                        'descripcion_raw': match_fmr.group(5),
                    }
                elif fmr_mode and current_fmr_item:
                    # Acumular la descripción
                    if not (linea_clean.startswith('COMMODITY CODE') or linea_clean.startswith('ITEM ')):
                        current_fmr_item['descripcion_raw'] += " " + linea_clean

            if current_fmr_item:
                items_extraidos.append(current_fmr_item)
                
            # Procesar raw descripciones de FMR
            for it in items_extraidos:
                if 'descripcion_raw' in it:
                    raw = it['descripcion_raw']
                    peso_match = re.search(r'Peso.*?([\d\.,]+)\s*kg', raw, re.IGNORECASE)
                    peso = peso_match.group(1).replace(',', '.') if peso_match else ""
                    
                    size_match = re.search(r'([\d\.,]+(?:/\d+)?\s*["\']|[\d\.,]+\s*IN)', raw, re.IGNORECASE)
                    size = size_match.group(1).strip() if size_match else ""
                    
                    desc = re.sub(r'Peso.*?([\d\.,]+)\s*kg', '', raw, flags=re.IGNORECASE)
                    desc = re.sub(r'aproximado\s*=', '', desc, flags=re.IGNORECASE).strip()
                    desc = re.sub(r'\s+', ' ', desc).strip()
                    
                    it['descripcion'] = desc[:255] if desc else f"Item {it['item_code']}"
                    if peso:
                        it['peso_unitario_kg'] = peso
                    if size:
                        it['size_code'] = size
                    del it['descripcion_raw']
            
            datos['items'] = items_extraidos
                    
            return JsonResponse({'success': True, 'datos': datos})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'No file sent'})


# ──────────────────────────────────────────────────────────────────────────────
# COTIZACIÓN — CRUD y PDF
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def cotizacion_list(request):
    cotizaciones = Cotizacion.objects.select_related('orden_compra').all()
    return render(request, 'core/cotizacion_list.html', {'cotizaciones': cotizaciones})


@login_required
def cotizacion_create(request):
    if request.method == 'POST':
        form = CotizacionForm(request.POST)
        if form.is_valid():
            cot = form.save()
            messages.success(request, f'Cotización N° {cot.numero_cotizacion} creada.')
            return redirect('cotizacion_detail', cotizacion_id=cot.id)
        else:
            messages.error(request, f'Error: {form.errors}')
    else:
        form = CotizacionForm()
    return render(request, 'core/cotizacion_form.html', {'form': form, 'titulo': 'Nueva Cotización'})


@login_required
def cotizacion_detail(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    items = cotizacion.items.all()
    item_form = ItemCotizacionForm()

    if request.method == 'POST' and 'agregar_item' in request.POST:
        item_form = ItemCotizacionForm(request.POST)
        if item_form.is_valid():
            item = item_form.save(commit=False)
            item.cotizacion = cotizacion
            item.save()
            messages.success(request, 'Ítem agregado.')
            return redirect('cotizacion_detail', cotizacion_id=cotizacion_id)
        else:
            messages.error(request, f'Error: {item_form.errors}')

    return render(request, 'core/cotizacion_detail.html', {
        'cotizacion': cotizacion,
        'items': items,
        'item_form': item_form,
    })


@login_required
def cotizacion_item_delete(request, cotizacion_id, item_id):
    item = get_object_or_404(ItemCotizacion, id=item_id, cotizacion_id=cotizacion_id)
    item.delete()
    messages.success(request, 'Ítem eliminado.')
    return redirect('cotizacion_detail', cotizacion_id=cotizacion_id)


@login_required
def cotizacion_pdf(request, cotizacion_id):
    """Genera el PDF de Cotización replicando el formato Bark."""
    cot = get_object_or_404(Cotizacion, id=cotizacion_id)
    items = cot.items.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Cotizacion_{cot.numero_cotizacion:05d}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter,
                            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    styles = getSampleStyleSheet()

    dark_blue = colors.HexColor('#0d1220')
    light_bg  = colors.HexColor('#f5f7fa')

    title_style = ParagraphStyle('TS', parent=styles['Normal'], fontName='Helvetica-Bold',
                                  fontSize=18, textColor=dark_blue, leading=22)
    sub_style   = ParagraphStyle('SS', parent=styles['Normal'], fontName='Helvetica-Bold',
                                  fontSize=9, textColor=colors.white)
    body_style  = ParagraphStyle('BS', parent=styles['Normal'], fontName='Helvetica',
                                  fontSize=8.5, leading=11, textColor=colors.HexColor('#333333'))
    bold_body   = ParagraphStyle('BB', parent=styles['Normal'], fontName='Helvetica-Bold',
                                  fontSize=8.5, leading=11)

    # ── Membrete ──────────────────────────────────────────────────────────────
    empresa_txt = (
        "<b>MAESTRANZA BARK SPA</b><br/>"
        "Giro: Maestranza y Fabricaciones Metálicas<br/>"
        "RUT: 77.XXX.XXX-X<br/>"
        "Dirección: Agustinas 1442, Calama<br/>"
        "Fono: +56 9 4016 0112 | administracion@maestranzabark.cl"
    )
    cot_txt = (
        f"<font color='#0d1220'><b>COTIZACIÓN</b></font><br/>"
        f"<b>N° {cot.numero_cotizacion:05d}</b>"
    )
    header_tbl = Table([[Paragraph(empresa_txt, body_style), Paragraph(cot_txt, title_style)]],
                        colWidths=[310, 220])
    header_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 6))

    # ── Recuadro Fecha/Válido/Cliente ─────────────────────────────────────────
    fecha_str  = cot.fecha.strftime('%d/%m/%Y') if cot.fecha else 'N/A'
    valido_str = cot.valido_hasta.strftime('%d/%m/%Y') if cot.valido_hasta else 'N/A'
    box_data = [
        [Paragraph('<b>Fecha</b>', bold_body), Paragraph(fecha_str, body_style),
         Paragraph('<b>Cotización N°</b>', bold_body), Paragraph(f'{cot.numero_cotizacion:05d}', body_style)],
        [Paragraph('<b>Válido hasta</b>', bold_body), Paragraph(valido_str, body_style),
         Paragraph('<b>Cliente ID</b>', bold_body), Paragraph(cot.cliente_id or '—', body_style)],
    ]
    box_tbl = Table(box_data, colWidths=[100, 120, 100, 210])
    box_tbl.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cccccc')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('BACKGROUND', (0,0), (-1,-1), light_bg),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(box_tbl)
    story.append(Spacer(1, 8))

    # ── Tabla Contacto ────────────────────────────────────────────────────────
    contact_hdr = Table([[Paragraph('CONTACTO', sub_style)]], colWidths=[530])
    contact_hdr.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), dark_blue), ('PADDING', (0,0), (-1,-1), 4)]))
    story.append(contact_hdr)
    contact_data = [
        [Paragraph('<b>Atte. a:</b>', bold_body), Paragraph(cot.contacto_nombre or '—', body_style),
         Paragraph('<b>Cargo:</b>', bold_body), Paragraph(cot.contacto_cargo or '—', body_style)],
    ]
    contact_tbl = Table(contact_data, colWidths=[70, 200, 60, 200])
    contact_tbl.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cccccc')),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(contact_tbl)
    story.append(Spacer(1, 6))

    # ── Datos del Cliente ─────────────────────────────────────────────────────
    client_hdr = Table([[Paragraph('DATOS DEL CLIENTE', sub_style)]], colWidths=[530])
    client_hdr.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), dark_blue), ('PADDING', (0,0), (-1,-1), 4)]))
    story.append(client_hdr)
    client_data = [
        [Paragraph('<b>Razón Social:</b>', bold_body), Paragraph(cot.razon_social or '—', body_style),
         Paragraph('<b>Giro:</b>', bold_body), Paragraph(cot.giro or '—', body_style)],
        [Paragraph('<b>RUT:</b>', bold_body), Paragraph(cot.rut_receptor or '—', body_style),
         Paragraph('<b>Ciudad:</b>', bold_body), Paragraph(cot.ciudad_receptor or '—', body_style)],
        [Paragraph('<b>Dirección:</b>', bold_body), Paragraph(cot.direccion_receptor or '—', body_style), '', ''],
    ]
    client_tbl = Table(client_data, colWidths=[80, 195, 60, 195])
    client_tbl.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cccccc')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#eeeeee')),
        ('PADDING', (0,0), (-1,-1), 5),
        ('SPAN', (1,2), (3,2)),
    ]))
    story.append(client_tbl)
    story.append(Spacer(1, 10))

    # ── Tabla de Ítems ────────────────────────────────────────────────────────
    items_hdr_style = ParagraphStyle('IH', parent=styles['Normal'], fontName='Helvetica-Bold',
                                      fontSize=8, textColor=colors.white)
    item_tbl_data = [[
        Paragraph('N°', items_hdr_style),
        Paragraph('Descripción', items_hdr_style),
        Paragraph('Valor kg', items_hdr_style),
        Paragraph('C/U', items_hdr_style),
        Paragraph('KG c/u', items_hdr_style),
        Paragraph('KG total', items_hdr_style),
        Paragraph('Valor', items_hdr_style),
    ]]
    for i, it in enumerate(items, 1):
        bg = colors.white if i % 2 == 1 else light_bg
        item_tbl_data.append([
            Paragraph(str(i), body_style),
            Paragraph(it.descripcion, body_style),
            Paragraph(f"${it.valor_kg:,.0f}", body_style),
            Paragraph(str(it.cantidad), body_style),
            Paragraph(f"{it.kg_por_unidad:,.3f}", body_style),
            Paragraph(f"{it.kg_total:,.3f}", body_style),
            Paragraph(f"${it.valor:,.0f}", body_style),
        ])
    if not items:
        item_tbl_data.append([Paragraph('Sin ítems.', body_style), '', '', '', '', '', ''])

    item_tbl = Table(item_tbl_data, colWidths=[25, 195, 65, 40, 60, 65, 80])
    item_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), dark_blue),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_bg]),
    ]))
    story.append(item_tbl)
    story.append(Spacer(1, 10))

    # ── Totales ────────────────────────────────────────────────────────────────
    totales_data = [
        [Paragraph('Sub Total', bold_body), Paragraph(f"${cot.subtotal:,.0f}", body_style)],
        [Paragraph('Impuesto 19%', bold_body), Paragraph(f"${cot.iva:,.0f}", body_style)],
        [Paragraph('<b>Total</b>', bold_body), Paragraph(f"<b>${cot.total:,.0f}</b>", bold_body)],
    ]
    totales_tbl = Table(totales_data, colWidths=[370, 160])
    totales_tbl.setStyle(TableStyle([
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('LINEABOVE', (0,2), (-1,2), 1, dark_blue),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(totales_tbl)

    if cot.observaciones:
        story.append(Spacer(1, 8))
        story.append(Paragraph(f'<b>Notas:</b> {cot.observaciones}', body_style))

    doc.build(story)
    return response


# ──────────────────────────────────────────────────────────────────────────────
# GUÍA DE DESPACHO — CRUD y PDF
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def guia_create(request, numero_oc, entrega_id):
    orden_compra = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    entrega = get_object_or_404(Entrega, id=entrega_id, orden_compra=orden_compra)

    if hasattr(entrega, 'guia_despacho_obj'):
        messages.info(request, 'Esta entrega ya tiene una Guía de Despacho.')
        return redirect('guia_detail', numero_oc=numero_oc, entrega_id=entrega_id)

    if request.method == 'POST':
        form = GuiaDespachoForm(request.POST)
        if form.is_valid():
            guia = form.save(commit=False)
            guia.entrega = entrega
            guia.save()
            registrar_trazabilidad(orden_compra, 'Guía de Despacho Creada',
                                   f'Se creó la Guía N° {guia.numero_guia}.', request.user)
            messages.success(request, f'Guía N° {guia.numero_guia} creada.')
            return redirect('guia_detail', numero_oc=numero_oc, entrega_id=entrega_id)
        else:
            messages.error(request, f'Error: {form.errors}')
    else:
        initial = {
            'numero_guia': entrega.guia_despacho or '',
            'fecha_emision': entrega.fecha_entrega,
            'receptor_nombre': orden_compra.cliente,
        }
        form = GuiaDespachoForm(initial=initial)

    return render(request, 'core/guia_form.html', {
        'form': form, 'orden_compra': orden_compra, 'entrega': entrega,
        'titulo': 'Crear Guía de Despacho',
    })


@login_required
def guia_detail(request, numero_oc, entrega_id):
    orden_compra = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    entrega = get_object_or_404(Entrega, id=entrega_id, orden_compra=orden_compra)
    guia = getattr(entrega, 'guia_despacho_obj', None)
    item_form = ItemGuiaForm()
    items = guia.items_guia.all() if guia else []

    if request.method == 'POST' and guia and 'agregar_item_guia' in request.POST:
        item_form = ItemGuiaForm(request.POST)
        if item_form.is_valid():
            ig = item_form.save(commit=False)
            ig.guia = guia
            ig.save()
            messages.success(request, 'Ítem agregado a la guía.')
            return redirect('guia_detail', numero_oc=numero_oc, entrega_id=entrega_id)

    return render(request, 'core/guia_detail.html', {
        'orden_compra': orden_compra, 'entrega': entrega,
        'guia': guia, 'items': items, 'item_form': item_form,
    })


@login_required
def guia_item_delete(request, numero_oc, entrega_id, item_id):
    item = get_object_or_404(ItemGuia, id=item_id)
    item.delete()
    messages.success(request, 'Ítem eliminado.')
    return redirect('guia_detail', numero_oc=numero_oc, entrega_id=entrega_id)


def _build_guia_story(guia, styles, body_style, bold_body, sub_style, items_hdr_style):
    """Construye la historia ReportLab de la Guía de Despacho (una o más páginas)."""
    dark_blue = colors.HexColor('#0d1220')
    light_bg  = colors.HexColor('#f5f7fa')
    story = []

    # Membrete
    empresa_txt = (
        "<b>MAESTRANZA BARK SPA</b><br/>"
        "Giro: Maestranza y Fabricaciones Metálicas<br/>"
        "RUT: 77.XXX.XXX-X<br/>"
        "Dirección: Agustinas 1442, Calama"
    )
    guia_txt = f"<b>GUÍA DE DESPACHO<br/>N° {guia.numero_guia}</b>"
    title_style = ParagraphStyle('GT', parent=styles['Normal'], fontName='Helvetica-Bold',
                                  fontSize=14, textColor=dark_blue, leading=18)
    hdr = Table([[Paragraph(empresa_txt, body_style), Paragraph(guia_txt, title_style)]],
                 colWidths=[310, 220])
    hdr.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('BOTTOMPADDING', (0,0), (-1,-1), 10)]))
    story.append(hdr)
    story.append(Spacer(1, 8))

    # Datos receptor
    oc_num = guia.entrega.orden_compra.numero_oc if guia.entrega else '—'
    receptor_data = [
        [Paragraph('<b>Fecha Emisión</b>', bold_body),
         Paragraph(guia.fecha_emision.strftime('%d/%m/%Y') if guia.fecha_emision else '—', body_style),
         Paragraph('<b>OC Referencia</b>', bold_body), Paragraph(oc_num, body_style)],
        [Paragraph('<b>Razón Social</b>', bold_body), Paragraph(guia.receptor_nombre or '—', body_style),
         Paragraph('<b>RUT</b>', bold_body), Paragraph(guia.receptor_rut or '—', body_style)],
        [Paragraph('<b>Giro</b>', bold_body), Paragraph(guia.receptor_giro or '—', body_style),
         Paragraph('<b>Comuna</b>', bold_body), Paragraph(guia.receptor_comuna or '—', body_style)],
        [Paragraph('<b>Dirección</b>', bold_body), Paragraph(guia.receptor_direccion or '—', body_style),
         Paragraph('<b>Contacto</b>', bold_body), Paragraph(guia.contacto or '—', body_style)],
    ]
    rec_tbl = Table(receptor_data, colWidths=[90, 190, 80, 170])
    rec_tbl.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cccccc')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#eeeeee')),
        ('BACKGROUND', (0,0), (-1,-1), light_bg),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(rec_tbl)
    story.append(Spacer(1, 6))

    # Datos transporte
    trans_data = [
        [Paragraph('<b>Tipo Despacho</b>', bold_body), Paragraph(guia.tipo_despacho or '—', body_style),
         Paragraph('<b>Tipo Traslado</b>', bold_body), Paragraph(guia.tipo_traslado or '—', body_style)],
        [Paragraph('<b>Chofer</b>', bold_body), Paragraph(guia.chofer_nombre or '—', body_style),
         Paragraph('<b>RUT Chofer</b>', bold_body), Paragraph(guia.chofer_rut or '—', body_style)],
        [Paragraph('<b>Patente</b>', bold_body), Paragraph(guia.patente or '—', body_style),
         Paragraph('<b>RUT Transportista</b>', bold_body), Paragraph(guia.transportista_rut or '—', body_style)],
        [Paragraph('<b>Dirección Destino</b>', bold_body), Paragraph(guia.direccion_destino or '—', body_style), '', ''],
    ]
    trans_tbl = Table(trans_data, colWidths=[100, 165, 100, 165])
    trans_tbl.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cccccc')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#eeeeee')),
        ('PADDING', (0,0), (-1,-1), 5),
        ('SPAN', (1,3), (3,3)),
    ]))
    story.append(trans_tbl)
    story.append(Spacer(1, 10))

    # Tabla detalle ítems
    items_tbl_data = [[
        Paragraph('N°', items_hdr_style),
        Paragraph('Descripción', items_hdr_style),
        Paragraph('Cantidad / Unidad', items_hdr_style),
        Paragraph('Precio Unit.', items_hdr_style),
        Paragraph('Total', items_hdr_style),
    ]]
    guia_items = list(guia.items_guia.all())
    for i, it in enumerate(guia_items, 1):
        items_tbl_data.append([
            Paragraph(str(i), body_style),
            Paragraph(it.descripcion, body_style),
            Paragraph(it.cantidad_unidad or '—', body_style),
            Paragraph(f"${it.precio_unitario:,.0f}", body_style),
            Paragraph(f"${it.total:,.0f}", body_style),
        ])
    if not guia_items:
        items_tbl_data.append([Paragraph('Sin ítems.', body_style), '', '', '', ''])

    it_tbl = Table(items_tbl_data, colWidths=[30, 260, 100, 80, 80])
    it_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), dark_blue),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_bg]),
    ]))
    story.append(it_tbl)
    story.append(Spacer(1, 8))

    # Totales
    tot_data = [
        [Paragraph('Monto Neto', bold_body), Paragraph(f"${guia.monto_neto:,.0f}", body_style)],
        [Paragraph('IVA (19%)', bold_body), Paragraph(f"${guia.iva:,.0f}", body_style)],
        [Paragraph('<b>Monto Total</b>', bold_body), Paragraph(f"<b>${guia.monto_total:,.0f}</b>", bold_body)],
    ]
    tot_tbl = Table(tot_data, colWidths=[420, 110])
    tot_tbl.setStyle(TableStyle([
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('LINEABOVE', (0,2), (-1,2), 1, dark_blue),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(tot_tbl)
    return story


@login_required
def guia_pdf(request, numero_oc, entrega_id):
    orden_compra = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    entrega = get_object_or_404(Entrega, id=entrega_id, orden_compra=orden_compra)
    guia = get_object_or_404(GuiaDespacho, entrega=entrega)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Guia_{guia.numero_guia}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter,
                            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle('BS', parent=styles['Normal'], fontName='Helvetica',
                                 fontSize=8.5, leading=11, textColor=colors.HexColor('#333333'))
    bold_body  = ParagraphStyle('BB', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.5, leading=11)
    sub_style  = ParagraphStyle('SbS', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
    items_hdr_style = ParagraphStyle('IH', parent=styles['Normal'], fontName='Helvetica-Bold',
                                      fontSize=8, textColor=colors.white)

    story = _build_guia_story(guia, styles, body_style, bold_body, sub_style, items_hdr_style)
    doc.build(story)
    return response


@login_required
def guia_packing_combinado_pdf(request, numero_oc, entrega_id):
    """PDF combinado: Guía de Despacho (pág 1) + Packing List (págs siguientes)."""
    from reportlab.platypus import PageBreak
    orden_compra = get_object_or_404(OrdenCompra, numero_oc=numero_oc)
    entrega = get_object_or_404(Entrega, id=entrega_id, orden_compra=orden_compra)
    guia = get_object_or_404(GuiaDespacho, entrega=entrega)

    # Obtener el primer packing list de esta entrega
    pl = entrega.packing_lists.first()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Guia_PL_{guia.numero_guia}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter,
                            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    dark_blue = colors.HexColor('#0d1220')
    light_bg  = colors.HexColor('#f5f7fa')

    body_style = ParagraphStyle('BS', parent=styles['Normal'], fontName='Helvetica',
                                 fontSize=8.5, leading=11, textColor=colors.HexColor('#333333'))
    bold_body  = ParagraphStyle('BB', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.5, leading=11)
    sub_style  = ParagraphStyle('SbS', parent=styles['Normal'], fontName='Helvetica-Bold',
                                  fontSize=9, textColor=colors.white)
    items_hdr_style = ParagraphStyle('IH', parent=styles['Normal'], fontName='Helvetica-Bold',
                                      fontSize=8, textColor=colors.white)
    title_style = ParagraphStyle('TS', parent=styles['Normal'], fontName='Helvetica-Bold',
                                  fontSize=15, textColor=dark_blue, alignment=1)
    header_style = ParagraphStyle('HS', parent=styles['Normal'], fontName='Helvetica-Bold',
                                   fontSize=9, leading=11, textColor=colors.white)

    story = _build_guia_story(guia, styles, body_style, bold_body, sub_style, items_hdr_style)
    story.append(PageBreak())

    # ── Packing List story ───────────────────────────────────────────────────
    if pl:
        empresa_info = (f"<b>{pl.empresa}</b><br/>Giro: Maestranza y Fabricaciones Metálicas<br/>"
                        f"Dir: {pl.direccion}<br/>Correo: {pl.correo}<br/>Fono: {pl.telefono}")
        documento_info = (f"<font color='#0D1220'>Packing List N° {pl.numero_correlativo:05d}</font><br/><br/>"
                          f"<b>CONTROL DE CALIDAD<br/>PACKING LIST</b>")
        pl_hdr = Table([[Paragraph(empresa_info, body_style), Paragraph(documento_info, title_style)]],
                        colWidths=[250, 280])
        pl_hdr.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('BOTTOMPADDING', (0,0), (-1,-1), 10)]))
        story.append(pl_hdr)
        story.append(Spacer(1, 10))

        cliente_data = [
            [Paragraph(f"<b>Cliente:</b> {pl.nombre_cliente}", body_style),
             Paragraph(f"<b>Fecha Orden:</b> {pl.fecha_orden.strftime('%d-%m-%Y') if pl.fecha_orden else 'N/A'}", body_style)],
            [Paragraph(f"<b>N° OC Asociada:</b> {pl.orden_compra.numero_oc}", body_style),
             Paragraph(f"<b>Fecha Envío:</b> {pl.fecha_envio.strftime('%d-%m-%Y') if pl.fecha_envio else 'N/A'}", body_style)],
            [Paragraph(f"<b>Guía Despacho:</b> {guia.numero_guia}", body_style),
             Paragraph("", body_style)],
        ]
        cl_tbl = Table(cliente_data, colWidths=[265, 265])
        cl_tbl.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#dddddd')),
            ('BACKGROUND', (0,0), (-1,-1), light_bg),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(cl_tbl)
        story.append(Spacer(1, 15))

        col_m1 = pl.col_medida_1
        col_m2 = pl.col_medida_2
        pl_items_data = [[
            Paragraph('Ítem / Descripción', header_style),
            Paragraph('Modelo Soporte', header_style),
            Paragraph(col_m1, header_style),
            Paragraph(col_m2, header_style),
            Paragraph('Estado', header_style),
            Paragraph('Unidades', header_style),
        ]]
        pl_items = list(entrega.packing_list_items.select_related('item_oc').all())
        for idx, pli in enumerate(pl_items, 1):
            m1 = str(pli.medida_1) if pli.medida_1 is not None else (pli.diametro or 'N/A')
            m2 = str(pli.medida_2) if pli.medida_2 is not None else (pli.alto_item or 'N/A')
            pl_items_data.append([
                Paragraph(f"{idx}. {pli.item_oc.descripcion}", body_style),
                Paragraph(pli.modelo_soporte or 'N/A', body_style),
                Paragraph(m1, body_style), Paragraph(m2, body_style),
                Paragraph(pli.estado_item or 'N/A', body_style),
                Paragraph(pli.unidades or str(int(pli.cantidad)), body_style),
            ])
        if not pl_items:
            pl_items_data.append([Paragraph('Sin ítems.', body_style), '', '', '', '', ''])

        pl_it_tbl = Table(pl_items_data, colWidths=[180, 110, 60, 60, 60, 60])
        pl_it_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), dark_blue),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_bg]),
        ]))
        story.append(pl_it_tbl)
    else:
        story.append(Paragraph('No hay Packing List asociado a esta entrega.', body_style))

    doc.build(story)
    return response
